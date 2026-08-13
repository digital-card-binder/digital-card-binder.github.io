from __future__ import annotations

# Source: the public [최종]팩전종수집 Google Sheet exported as XLSX.
import argparse
import math
import statistics
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageFilter

PACK_ORDER = [
    "s1W", "s1H", "s1a", "s2", "s2a", "s3", "s3a", "s4", "s4a",
    "s5I", "s5R", "s5a", "s6H", "s6K", "s6a", "s7D", "s7R", "s8",
    "s8a", "s8b", "s9", "s9a", "s10P", "s10D", "s10a", "s10b",
    "s11", "s11a", "s12", "s12a", "sv1S", "sv1V", "sv1a", "sv2D",
    "sv2P", "sv2a", "sv3", "sv3a", "sv4K", "sv4M", "sv4a", "sv5K",
    "sv5M", "sv5a", "sv6", "sv6a", "sv7", "sv7a", "sv8", "sv8a",
    "sv9", "sv9a", "sv10", "sv11B", "sv11W", "m1S", "m1L", "m2",
    "m2a", "m3", "m4", "m5",
]


def edge_background(image: Image.Image) -> tuple[int, int, int, int]:
    width, height = image.size
    coords: list[tuple[int, int]] = []
    for x in (0, 1, max(0, width - 2), max(0, width - 1)):
        coords.extend((x, min(y, height - 1)) for y in range(0, height, max(1, height // 20)))
    for y in (0, 1, max(0, height - 2), max(0, height - 1)):
        coords.extend((min(x, width - 1), y) for x in range(0, width, max(1, width // 20)))
    values = [image.getpixel(point) for point in coords]
    return tuple(int(statistics.median(value[channel] for value in values)) for channel in range(4))


def crop_pack(image: Image.Image) -> Image.Image:
    image = image.convert("RGBA")
    background = edge_background(image)
    if background[3] < 100:
        mask = image.getchannel("A").point(lambda value: 255 if value > 20 else 0)
    else:
        flat = Image.new("RGBA", image.size, background)
        mask = ImageChops.difference(image, flat).convert("L").point(
            lambda value: 255 if value > 28 else 0
        )
    mask = mask.filter(ImageFilter.MaxFilter(15))
    box = mask.getbbox() or image.getbbox() or (0, 0, image.width, image.height)
    padding = max(3, int(min(image.size) * 0.015))
    left = max(0, box[0] - padding)
    top = max(0, box[1] - padding)
    right = min(image.width, box[2] + padding)
    bottom = min(image.height, box[3] + padding)
    return image.crop((left, top, right, bottom))


def extract_images(source: Path) -> dict[str, Image.Image]:
    workbook = load_workbook(source)
    images: dict[str, Image.Image] = {}
    for sheet in workbook.worksheets:
        for embedded in sheet._images:
            row = embedded.anchor._from.row + 1
            column = embedded.anchor._from.col + 1
            label = None
            for candidate_row in range(row, min(sheet.max_row, row + 3) + 1):
                value = sheet.cell(candidate_row, column).value
                if isinstance(value, str) and "\n" in value:
                    label = value
                    break
            if label is None:
                raise RuntimeError(f"Could not match image at {sheet.title}!R{row}C{column}")
            name, code = label.rsplit("\n", 1)
            if "화이트플레어" in name:
                code = "sv11W"
            images[code] = crop_pack(Image.open(BytesIO(embedded._data())))
    missing = [code for code in PACK_ORDER if code not in images]
    if missing:
        raise RuntimeError(f"Missing pack images: {', '.join(missing)}")
    if len(images) != len(PACK_ORDER):
        raise RuntimeError(f"Expected {len(PACK_ORDER)} images, found {len(images)}")
    return images


def build_sprite(images: dict[str, Image.Image], destination: Path) -> None:
    cell_width, cell_height, columns = 180, 230, 10
    rows = math.ceil(len(PACK_ORDER) / columns)
    sprite = Image.new("RGBA", (cell_width * columns, cell_height * rows), (0, 0, 0, 0))
    for index, code in enumerate(PACK_ORDER):
        image = images[code].copy().convert("RGBA")
        image.thumbnail((cell_width - 18, cell_height - 18), Image.Resampling.LANCZOS)
        x = (index % columns) * cell_width + (cell_width - image.width) // 2
        y = (index // columns) * cell_height + (cell_height - image.height) // 2
        sprite.alpha_composite(image, (x, y))
    destination.parent.mkdir(parents=True, exist_ok=True)
    sprite.save(destination, "WEBP", quality=84, method=6)
    print(f"Wrote {destination} ({destination.stat().st_size:,} bytes, {len(PACK_ORDER)} packs)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the pack collection WebP sprite from the public Google Sheet export.")
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    build_sprite(extract_images(args.source), args.destination)


if __name__ == "__main__":
    main()
