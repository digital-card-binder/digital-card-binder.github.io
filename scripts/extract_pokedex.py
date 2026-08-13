from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


GENERATION_SHEETS = [f"{number}세대" for number in range(1, 10)]
IMAGE_URL_RE = re.compile(r'IMAGE\("([^"]+)"', re.IGNORECASE)


def extract_image_url(value: object) -> str:
    match = IMAGE_URL_RE.search(str(value or ""))
    return match.group(1) if match else ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract the 1,025 Pokémon national dex from the source workbook."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = load_workbook(
        args.workbook,
        read_only=False,
        data_only=False,
        keep_vba=True,
        keep_links=True,
    )

    records: list[dict[str, object]] = []
    generation_summaries: list[dict[str, object]] = []

    for generation, sheet_name in enumerate(GENERATION_SHEETS, start=1):
        sheet = workbook[sheet_name]
        generation_records: list[dict[str, object]] = []

        for row_number in range(3, sheet.max_row + 1):
            raw_number = sheet.cell(row_number, 1).value
            if raw_number in (None, ""):
                continue

            number = int(str(raw_number).lstrip("#"))
            raw_status = sheet.cell(row_number, 5).value
            if raw_status not in {"☑", "☐"}:
                raise ValueError(
                    f"Unexpected ownership status in {sheet_name}!E{row_number}: "
                    f"{raw_status!r}"
                )

            record = {
                "number": number,
                "numberLabel": f"#{number:04d}",
                "generation": generation,
                "nameKo": str(sheet.cell(row_number, 4).value or "").strip(),
                "nameEn": str(sheet.cell(row_number, 3).value or "").strip(),
                "imageUrl": extract_image_url(sheet.cell(row_number, 2).value),
                "owned": raw_status == "☑",
            }
            generation_records.append(record)
            records.append(record)

        owned = sum(bool(record["owned"]) for record in generation_records)
        generation_summaries.append(
            {
                "generation": generation,
                "count": len(generation_records),
                "owned": owned,
                "missing": len(generation_records) - owned,
                "completionRate": round(
                    owned / len(generation_records) * 100, 1
                ),
                "firstNumber": generation_records[0]["number"],
                "lastNumber": generation_records[-1]["number"],
            }
        )

    numbers = [int(record["number"]) for record in records]
    number_counts = Counter(numbers)
    duplicate_numbers = sorted(
        number for number, count in number_counts.items() if count > 1
    )
    missing_numbers = sorted(set(range(1, 1026)) - set(numbers))
    incomplete_records = [
        record["number"]
        for record in records
        if not all(
            [
                record["nameKo"],
                record["nameEn"],
                record["imageUrl"],
            ]
        )
    ]

    if len(records) != 1025:
        raise ValueError(f"Expected 1,025 Pokémon, found {len(records)}")
    if duplicate_numbers:
        raise ValueError(f"Duplicate Pokédex numbers: {duplicate_numbers}")
    if missing_numbers:
        raise ValueError(f"Missing Pokédex numbers: {missing_numbers}")
    if incomplete_records:
        raise ValueError(f"Incomplete records: {incomplete_records}")

    owned = sum(bool(record["owned"]) for record in records)
    payload = {
        "meta": {
            "title": "1025 전국도감",
            "recordCount": len(records),
            "owned": owned,
            "missing": len(records) - owned,
            "completionRate": round(owned / len(records) * 100, 1),
            "generatedOn": date.today().isoformat(),
            "sourceFile": "[최종]1025전국도감.xlsm",
        },
        "generations": generation_summaries,
        "records": records,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "output": str(args.output),
                "records": len(records),
                "owned": owned,
                "missing": len(records) - owned,
                "completionRate": payload["meta"]["completionRate"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
